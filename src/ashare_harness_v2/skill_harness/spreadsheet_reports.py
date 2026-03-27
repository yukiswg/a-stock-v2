from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils import ensure_dir


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(bottom=Side(style="thin", color="9E9E9E"))
NEGATIVE_FONT = Font(color="C00000")
POSITIVE_FONT = Font(color="006100")


def build_portfolio_candidate_workbook(
    config: dict[str, Any],
    *,
    as_of: str,
    market_view: dict[str, Any],
    holdings_rows: list[dict[str, Any]],
    candidate_rows: list[dict[str, Any]],
) -> dict[str, str]:
    root = ensure_dir(Path(config["project"].get("spreadsheet_dir") or "data/output/spreadsheet") / as_of)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    holdings = workbook.create_sheet("Holdings")
    candidates = workbook.create_sheet("Candidates")

    _write_overview(overview, as_of=as_of, market_view=market_view, holdings_count=len(holdings_rows), candidate_count=len(candidate_rows))
    _write_holdings(holdings, rows=holdings_rows)
    _write_candidates(candidates, rows=candidate_rows)

    path = root / f"{as_of}_portfolio_candidates.xlsx"
    workbook.save(path)
    return {"workbook": str(path)}


def _write_overview(sheet, *, as_of: str, market_view: dict[str, Any], holdings_count: int, candidate_count: int) -> None:
    sheet["A1"] = "A-share Portfolio Comparison"
    sheet["A1"].font = Font(size=16, bold=True)
    sheet["A3"] = "As Of"
    sheet["B3"] = as_of
    sheet["A4"] = "Market Label"
    sheet["B4"] = str((market_view.get("metadata") or {}).get("label") or market_view.get("action") or "")
    sheet["A5"] = "Market Score"
    sheet["B5"] = float(market_view.get("score") or 0.0) / 100.0
    sheet["B5"].number_format = "0.0%"
    sheet["A6"] = "Holdings Count"
    sheet["B6"] = holdings_count
    sheet["A7"] = "Candidates Count"
    sheet["B7"] = candidate_count
    sheet["A9"] = "Decision Notes"
    sheet["A10"] = "Holdings and candidates below share one scoring core. Use this sheet for relative ranking, not single-name conviction only."
    for cell in ("A3", "A4", "A5", "A6", "A7", "A9"):
        sheet[cell].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["A"].alignment = Alignment(vertical="top")
    sheet.column_dimensions["B"].alignment = Alignment(vertical="top")


def _write_holdings(sheet, *, rows: list[dict[str, Any]]) -> None:
    headers = [
        "Code",
        "Name",
        "Action",
        "Decision",
        "Total Score",
        "Market Score",
        "Coverage",
        "Weight",
        "Sector",
        "PnL",
        "Top Positive",
        "Top Negative",
    ]
    _write_header_row(sheet, headers)
    for index, row in enumerate(rows, start=2):
        scorecard = row.get("scorecard") or {}
        position_context = row.get("position_context") or {}
        sheet.cell(index, 1, row.get("code"))
        sheet.cell(index, 2, row.get("name"))
        sheet.cell(index, 3, row.get("action"))
        sheet.cell(index, 4, row.get("decision"))
        sheet.cell(index, 5, float(scorecard.get("total_score") or 0.0))
        sheet.cell(index, 6, float(scorecard.get("market_score") or 0.0))
        sheet.cell(index, 7, float(scorecard.get("coverage_score") or 0.0))
        sheet.cell(index, 8, float(position_context.get("position_weight") or 0.0))
        sheet.cell(index, 9, row.get("sector"))
        sheet.cell(index, 10, position_context.get("pnl_pct"))
        sheet.cell(index, 11, " / ".join((row.get("positive_factors") or [])[:2]))
        sheet.cell(index, 12, " / ".join((row.get("negative_factors") or [])[:2]))
        sheet.cell(index, 1).comment = Comment(str(row.get("summary") or ""), "Codex")
        sheet.cell(index, 8).number_format = "0.0%"
        sheet.cell(index, 10).number_format = "0.0%"
        if str(row.get("decision")) in {"add", "buy"}:
            sheet.cell(index, 4).font = POSITIVE_FONT
        elif str(row.get("decision")) in {"trim", "avoid"}:
            sheet.cell(index, 4).font = NEGATIVE_FONT
    _autosize(sheet)


def _write_candidates(sheet, *, rows: list[dict[str, Any]]) -> None:
    headers = [
        "Code",
        "Name",
        "Decision",
        "Total Score",
        "Coverage",
        "Market Score",
        "Sector",
        "Catalysts",
        "Risks",
        "Summary",
    ]
    _write_header_row(sheet, headers)
    for index, row in enumerate(rows, start=2):
        sheet.cell(index, 1, row.get("code"))
        sheet.cell(index, 2, row.get("name"))
        sheet.cell(index, 3, row.get("decision"))
        sheet.cell(index, 4, float(row.get("total_score") or 0.0))
        sheet.cell(index, 5, float(row.get("coverage_score") or 0.0))
        sheet.cell(index, 6, float(row.get("market_score") or 0.0))
        sheet.cell(index, 7, (row.get("metadata") or {}).get("sector"))
        sheet.cell(index, 8, " / ".join((row.get("catalysts") or [])[:3]))
        sheet.cell(index, 9, " / ".join((row.get("risks") or [])[:3]))
        sheet.cell(index, 10, row.get("summary"))
        if str(row.get("decision")) == "buy":
            sheet.cell(index, 3).font = POSITIVE_FONT
        elif str(row.get("decision")) in {"avoid", "watch"}:
            sheet.cell(index, 3).font = NEGATIVE_FONT if str(row.get("decision")) == "avoid" else Font(color="9C6500")
    _autosize(sheet)


def _write_header_row(sheet, headers: list[str]) -> None:
    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(1, column, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    sheet.freeze_panes = "A2"


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
